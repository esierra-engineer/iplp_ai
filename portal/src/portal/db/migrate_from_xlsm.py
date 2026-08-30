"""One-time importer: openpyxl (+, when available, the current golden .dat files) -> SQLite, for a
new Case.

`dat_static_dir`/`dat_block_dependant_dir` are optional. Passing them lets this importer bootstrap
the handful of fields that have no Excel/VBA source at all (see below) from the case's current
authoritative .dat output. Omitting either (or both) — the "xlsm alone" path — still imports
everything derivable from the workbook; the fields below fall back to sensible model defaults (or,
for tables with genuinely no Excel source, are simply left empty) instead of crashing, and each
fallback logs a warning so it's visible rather than silently assumed correct.

Phase 1 scope: buses (Barras), lines (Líneas), the stage calendar (Etapas), and the three
solver-control files (plpmat.dat/plpdeb.dat/plprun.dat).

Known bootstrap limitations (documented, not hidden):

- Stage.hydro_dependent (FDesh) and Stage.rate_factor (FactTasa) are not plain Etapas-sheet
  columns — FactTasa in particular is a compounding per-stage discount factor. This importer reads
  both from the case's existing dat/block_dependant/plpeta.dat when given (the current,
  authoritative output for this case) rather than re-deriving them, since that derivation isn't
  ported yet. With no golden file, they default to hydro_dependent=False, rate_factor=1.0.
- Block durations (plpblo.dat) ARE fully derived from the workbook: the Etapas sheet carries one
  column per block (column 8 onward) with that block's hour count directly, keyed off "Nº Bloques"
  (column 7) — no golden file needed for this one; see `_import_blocks`.
- plpmat.dat/plpdeb.dat/plprun.dat aren't clearly Excel-sourced at all (no sheet reference was found
  for them in the VBA map) — read from dat/static/ when given, else every column uses the model's
  own default (see db/models.py's MathParams/DebugParams/RunParams).

These are legitimate for a *migration* importer (its job is exactly "seed the DB from whatever is
currently authoritative for this case"), and the defaults above make a from-scratch new case
(created from only an .xlsm, with no pre-existing .dat files at all) still produce a usable case —
those specific fields just start at neutral defaults, editable afterwards via the web UI, rather
than something derived from a ported VBA algorithm that doesn't exist yet.

Phase 2 (plant fleet) adds a similar situation, worth calling out separately since it's not a
scoping shortcut but a genuine absence: **no VBA writer for plpcenre.dat or plpcenpmax.dat was
found in either xla workbook** (searched both). Those two files' reservoir rating-curve data has no
Excel source at all — bootstrapped from the golden files when given, otherwise left empty (no
plausible default exists — these are per-reservoir piecewise curves, not scalars). Separately,
plpcnfce.dat's per-embalse `EmbFEsc` (scale factor) also has no Centrales-sheet column (confirmed by
cross-checking real values against the parsed golden file) and is bootstrapped the same way when a
golden file is given, else defaults to 1.0 (a neutral scale factor); every other plpcnfce.dat field
below (including the header's 5 constant flags, and 9 per-plant fields confirmed uniform-constant
across all 2964 plants in this case: cen_ipot, min_tec, inter, fcad, mttd_hrz, cost_arranque,
cost_detencion, on_flag, p_ini) comes straight from the Centrales sheet — see `_import_plants`'s
column-mapping comment, empirically validated against the golden file's actual values (not just the
sheet's header labels, which are ambiguous in a couple of spots).

Phases 4-6's maintenance/hydrology/basin-convention tables (plpmanbat.dat, plpaflce.dat/
plpidsim.dat/plpidape.dat/plpidap2.dat, plpralco.dat/plpextrac.dat/plpfilemb.dat/plpvrebemb.dat)
have no Excel source at all regardless of xlsm-alone status (see each function's docstring) — they
already guard each golden file's existence individually and simply import nothing (empty tables,
not an error) when it's unavailable.
"""

from __future__ import annotations

from datetime import timedelta
from pathlib import Path

import openpyxl
from sqlalchemy import insert, select
from sqlalchemy.orm import Session

from ..curves.reservoir_volume import volume_from_level
from ..curves.reservoir_yield import yield_from_level
from ..dat_readers import (
    parse_lines_raw,
    parse_plpaflce,
    parse_plpblo,
    parse_plpcenbat,
    parse_plpcenpmax,
    parse_plpcenre,
    parse_plpcnfce,
    parse_plpdeb,
    parse_plpeta,
    parse_plpextrac,
    parse_plpfilemb,
    parse_plpidap2,
    parse_plpidape,
    parse_plpidsim,
    parse_plpmanbat,
    parse_plpmat,
    parse_plpralco,
    parse_plprun,
    parse_plpvrebemb,
)
from .models import (
    ApertureIndexAggregate,
    ApertureIndexSimulation,
    BasinConventionLine,
    Battery,
    BatteryInjector,
    BatteryMaintenance,
    ExtractionPoint,
    Bus,
    Case,
    ConsumptionWeek,
    DebugParams,
    DemandProfile,
    Holiday,
    HydrologyScenarioAssignment,
    IndustrialProject,
    Inflow,
    Line,
    LineConfig,
    LineMaintenance,
    MathParams,
    Plant,
    PlantMaintenance,
    RalcoConvention,
    Reservoir,
    ReservoirFiltration,
    ReservoirMaintenance,
    ReservoirMinVolumeSlack,
    ReservoirPmaxCurve,
    ReservoirPmaxSegment,
    ReservoirSpillVolume,
    ReservoirYieldCurve,
    ReservoirYieldSegment,
    RunParams,
    Stage,
    Block,
    ThermalCostSchedule,
)

# Centrales sheet "Tipo de Central" single-letter code -> Plant.plant_type. 'E'/'A' and 'S'/'R' each
# collapse to one plant_type; which sub-code a plant gets at generation time is derived from
# bus_id (see generators/plpcnfce.py), matching leecnfce.f's own CenGBar==0 classification exactly.
# 'X' (fuera de servicio) plants are not imported at all — matches the VBA writer's own count.
_CENTRALES_TYPE_MAP = {
    "E": "EMBALSE",
    "S": "SERIE",
    "R": "SERIE",
    "P": "PASADA",
    "T": "TERMICA",
    "BAT": "BATERIA",
    "F": "FALLA",
}

_EXCEL_ERROR_STRINGS = {"#NAME?", "#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"}


def _safe_float(value, *, context: str, default: float = 0.0) -> float:
    """Convert a cell value to float, tolerating live Excel formula errors (a broken/unresolved
    reference cached as e.g. '#NAME?') rather than crashing the whole import on one bad cell —
    confirmed to happen in real, currently-in-use workbooks, not a hypothetical. Logs a warning
    (not silent) and falls back to `default` so the rest of the case still imports; the affected
    field can be corrected by hand afterwards via the web UI."""
    if value is None:
        return default
    if isinstance(value, str) and value.strip().upper() in _EXCEL_ERROR_STRINGS:
        print(f"import_case: warning, {context} is an Excel error ({value!r}) — using {default}.")
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        print(f"import_case: warning, {context} is not numeric ({value!r}) — using {default}.")
        return default


def import_case(
    session: Session,
    *,
    case_name: str,
    xlsm_path: Path,
    dat_static_dir: Path | None = None,
    dat_block_dependant_dir: Path | None = None,
    description: str | None = None,
    case_id: int | None = None,
) -> Case:
    """`case_id`, when given, is used as the new Case row's explicit primary key instead of letting
    SQLite autoincrement assign one — needed when this case gets its own dedicated SQLite file (see
    db/registry.py) and must carry the same id the registry already allocated it, so URLs/generator
    calls elsewhere don't need to know that file's own internal id separately."""
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=False)

    case = Case(id=case_id, name=case_name, description=description) if case_id is not None else Case(
        name=case_name, description=description
    )
    session.add(case)
    session.flush()  # assign case.id

    bus_by_num = _import_buses(session, case, wb)
    stage_by_num = _import_stages(session, case, wb, dat_block_dependant_dir)
    _import_blocks(session, case, wb, stage_by_num)
    _import_lines(session, case, bus_by_num, wb)
    _import_solver_params(session, case, dat_static_dir)

    plant_by_name = _import_plants(session, case, bus_by_num, wb, dat_static_dir)
    _import_reservoir_curves(session, case, plant_by_name, dat_static_dir)
    _import_batteries(session, case, bus_by_num, plant_by_name, wb)

    bus_by_upper_name = {b.name.upper(): b for b in bus_by_num.values()}
    _import_demand_profiles(session, case, bus_by_upper_name, wb)
    _import_consumption_and_holidays(session, case, wb)
    _import_industrial_projects(session, case, bus_by_upper_name, wb)
    _import_thermal_cost_schedule(session, case, plant_by_name, wb)

    stages = session.scalars(select(Stage).where(Stage.case_id == case.id)).all()
    _import_plant_maintenance(session, case, plant_by_name, wb)
    _import_line_maintenance(session, case, wb)
    _import_reservoir_maintenance(session, case, plant_by_name, wb)
    _import_reservoir_min_volume_slack(session, case, plant_by_name, stages, wb)
    _import_battery_maintenance(session, case, dat_block_dependant_dir)

    stage_by_num = {s.num_eta: s for s in stages}
    _import_hydrology(session, case, plant_by_name, stage_by_num, dat_block_dependant_dir)

    _import_basin_conventions(session, case, plant_by_name, dat_static_dir)

    return case


def _import_buses(session: Session, case: Case, wb) -> dict[int, Bus]:
    """Barras sheet: header row 5 ('Nº', 'BARRA'), data from row 6."""
    ws = wb["Barras"]
    bus_by_num: dict[int, Bus] = {}
    row = 6
    while True:
        num = ws.cell(row, 1).value
        name = ws.cell(row, 2).value
        if num is None or name is None:
            break
        bus = Bus(case_id=case.id, num_bar=int(num), name=str(name))
        session.add(bus)
        bus_by_num[int(num)] = bus
        row += 1
    session.flush()
    return bus_by_num


def _import_stages(
    session: Session, case: Case, wb, dat_block_dependant_dir: Path | None
) -> dict[int, Stage]:
    """Etapas sheet: header row 4 ('Etapa','Inicial','Nº Días','Final','Año','Mes','Nº Bloques'),
    data from row 5. FDesh/FactTasa are bootstrapped from the golden plpeta.dat when available (see
    module docstring); with no golden file they default to hydro_dependent=False, rate_factor=1.0."""
    ws = wb["Etapas"]
    golden_path = dat_block_dependant_dir / "plpeta.dat" if dat_block_dependant_dir else None
    if golden_path is not None and golden_path.exists():
        golden = parse_plpeta(golden_path.read_text(encoding="latin-1"))
        golden_by_num = {s["num_eta"]: s for s in golden["stages"]}
    else:
        print(
            "import_case: warning, no plpeta.dat golden file — Stage.hydro_dependent/rate_factor "
            "default to False/1.0."
        )
        golden_by_num = {}

    stage_by_num: dict[int, Stage] = {}
    row = 5
    while True:
        num_eta = ws.cell(row, 1).value
        inicial = ws.cell(row, 2).value
        n_dias = ws.cell(row, 3).value
        n_bloques = ws.cell(row, 7).value
        if num_eta is None or inicial is None:
            break
        num_eta = int(num_eta)
        g = golden_by_num.get(num_eta, {})
        stage = Stage(
            case_id=case.id,
            num_eta=num_eta,
            year=inicial.year,
            month=inicial.month,
            hydro_dependent=g.get("hydro_dependent", False),
            duration=int(n_dias) * 24,
            rate_factor=g.get("rate_factor", 1.0),
            label=f"{int(n_bloques)} Bloques",
            start_date=inicial.date(),
        )
        session.add(stage)
        stage_by_num[num_eta] = stage
        row += 1
    session.flush()
    return stage_by_num


def _import_blocks(session: Session, case: Case, wb, stage_by_num: dict[int, Stage]) -> None:
    """plpblo.dat block durations: derived directly from the Etapas sheet itself, not bootstrapped
    from a golden file — 'Nº Bloques' (column 7) says how many blocks a stage has, and columns 8
    onward carry that many per-block hour counts, one column per block (confirmed against the
    golden plpblo.dat: e.g. stage 1's 10 values in columns 8-17 match its 10 'NHoras' rows exactly).
    num_blo is a global sequential counter across every stage, matching the golden file's own
    strictly-ascending 1..N numbering."""
    ws = wb["Etapas"]
    num_blo = 0
    row = 5
    while True:
        num_eta = ws.cell(row, 1).value
        if num_eta is None:
            break
        num_eta = int(num_eta)
        stage = stage_by_num[num_eta]
        n_bloques = int(ws.cell(row, 7).value)
        for i in range(n_bloques):
            num_blo += 1
            hours = ws.cell(row, 8 + i).value
            session.add(
                Block(
                    case_id=case.id,
                    num_blo=num_blo,
                    stage_id=stage.id,
                    duration=int(
                        _safe_float(hours, context=f"Etapas row {row}, block {i + 1} hours", default=0.0)
                    ),
                    year=stage.year,
                    month=stage.month,
                    label=f"Bloque {num_blo:02d}",
                )
            )
        row += 1
    session.flush()


def _import_lines(session: Session, case: Case, bus_by_num: dict[int, Bus], wb) -> None:
    """Líneas sheet: header row 5, data from row 6 (see column mapping in generators/plpcnfli.py)."""
    ws = wb["Líneas"]
    session.add(
        LineConfig(
            case_id=case.id,
            models_losses_globally=bool(ws.cell(1, 13).value),
            loss_model_in_erm=str(ws.cell(2, 13).value),
        )
    )
    row = 6
    while True:
        name = ws.cell(row, 2).value
        if name is None:
            break
        bus_a = bus_by_num[int(ws.cell(row, 3).value)]
        bus_b = bus_by_num[int(ws.cell(row, 4).value)]
        session.add(
            Line(
                case_id=case.id,
                name=str(name),
                bus_from_id=bus_a.id,
                bus_to_id=bus_b.id,
                capacity_ab=_safe_float(ws.cell(row, 5).value, context=f"Líneas!{name}, capacity_ab"),
                capacity_ba=_safe_float(ws.cell(row, 6).value, context=f"Líneas!{name}, capacity_ba"),
                voltage_kv=_safe_float(ws.cell(row, 7).value, context=f"Líneas!{name}, voltage_kv"),
                resistance=_safe_float(ws.cell(row, 10).value, context=f"Líneas!{name}, resistance"),
                reactance=_safe_float(ws.cell(row, 11).value, context=f"Líneas!{name}, reactance"),
                models_losses=bool(ws.cell(row, 12).value),
                num_segments=int(ws.cell(row, 13).value),
                operational=bool(ws.cell(row, 14).value),
                is_hvdc=bool(ws.cell(row, 15).value) if ws.cell(row, 15).value is not None else None,
            )
        )
        row += 1
    session.flush()


def _import_solver_params(session: Session, case: Case, dat_static_dir: Path | None) -> None:
    """plpmat.dat/plpdeb.dat/plprun.dat: no confirmed Excel source (see module docstring) —
    bootstrapped from the golden static files when available; otherwise each column falls back to
    the model's own default (see db/models.py's MathParams/DebugParams/RunParams)."""

    def _load(parser, filename: str, label: str) -> dict:
        path = dat_static_dir / filename if dat_static_dir else None
        if path is not None and path.exists():
            return parser(path.read_text(encoding="latin-1"))
        print(f"import_case: warning, no {filename} golden file — {label} uses model defaults.")
        return {}

    mat = _load(parse_plpmat, "plpmat.dat", "MathParams")
    deb = _load(parse_plpdeb, "plpdeb.dat", "DebugParams")
    run = _load(parse_plprun, "plprun.dat", "RunParams")
    session.add(MathParams(case_id=case.id, **mat))
    session.add(DebugParams(case_id=case.id, **deb))
    session.add(RunParams(case_id=case.id, **run))


def _import_plants(
    session: Session, case: Case, bus_by_num: dict[int, Bus], wb, dat_static_dir: Path | None
) -> dict[str, Plant]:
    """Centrales sheet: header rows 4-5, data from row 6. Column mapping below was derived from the
    sheet's own header labels, then EMPIRICALLY VALIDATED by cross-checking real values (cen_ind,
    name, bus, downstream references, volumes, etc.) for one plant of each type against the parsed
    golden plpcnfce.dat — not assumed correct from the labels alone, which are ambiguous for the
    embalse-only volume columns (two similarly-named column groups exist; only one is the one
    leecnfce.f actually reads). See the plan/session notes for the full trace.

    col 1  = cen_ind ("INDICE")              col 9  = cfue ("Función Costo Futuro", Embalse only)
    col 2  = name                            col 12 = cau_afl ("Afluente Primera Semana")
    col 3  = type letter                     col 13 = estoc_findep ("Independencia Hidrológica")
    col 4  = cos_var ("Costo Variable")      col 23 = vol_ini ("Volumen del embalse hm3" Inicial)
    col 5  = rendimiento                     col 24 = vol_fin (... Final)
    col 6  = bus ("Conectada a la Barra")    col 25 = vol_min (... Mínimo)
    col 7  = gen_hid ref ("Generación")      col 26 = vol_max (... Máximo)
    col 8  = vert_hid ref ("Vertimiento")    col 27 = pot_min ("Potencia Neta" Mínima)
                                              col 28 = pot_max (... Máxima)
                                              col 29 = vert_min ("Vertimiento" Mínimo)
                                              col 30 = vert_max (... Máximo)

    (col 10, "Afluente Estocástico", tracks col 13 exactly in every sample checked — could be a
    duplicate/legacy column; col 13's label is the closer semantic match to Fortran's EstocFIndep
    and is what's used here.)

    gen_hid/vert_hid are the referenced plant's cen_ind (0/blank = no downstream) — resolved to
    Plant FKs in a second pass below since a plant can reference one that appears later in the sheet.

    Reservoir.vol_ini/vol_fin/vol_min/vol_max are stored here in the sheet's own units (hm3, e.g.
    LMAULE's ~757) — the .dat file token is `sheet_value * 1e6 / f_esc` (confirmed empirically
    against two embalses with different f_esc; see generators/plpcnfce.py for the formula and why
    f_esc's effect cancels out internally regardless). Storing raw sheet units in the DB (rather
    than the file-scaled value) keeps the web UI showing the same numbers an analyst already knows
    from Excel — the scaling is purely a file-writing convention, applied only at generation time.
    """
    ws = wb["Centrales"]
    golden_femb: dict[str, float] = {}
    cnfce_path = dat_static_dir / "plpcnfce.dat" if dat_static_dir else None
    if cnfce_path is not None and cnfce_path.exists():
        golden_femb = {
            p["name"]: p["f_esc"]
            for p in parse_plpcnfce(cnfce_path.read_text(encoding="latin-1"))["plants"]
            if p["block"] == "EMBALSE"
        }
    else:
        print("import_case: warning, no plpcnfce.dat golden file — Reservoir.f_esc defaults to 1.0.")

    plant_by_name: dict[str, Plant] = {}
    plant_by_cen_ind: dict[int, Plant] = {}
    pending_refs: list[tuple[Plant, int, int]] = []  # (plant, gen_hid_ref, vert_hid_ref)

    row = 6
    while True:
        cen_ind = ws.cell(row, 1).value
        if cen_ind is None:
            break
        type_letter = str(ws.cell(row, 3).value or "").strip().upper()
        plant_type = _CENTRALES_TYPE_MAP.get(type_letter)
        if plant_type is None:  # 'X' (fuera de servicio) or unrecognized — not part of the .dat file
            row += 1
            continue

        bus_num = ws.cell(row, 6).value
        bus = bus_by_num.get(int(bus_num)) if bus_num else None
        plant_name = str(ws.cell(row, 2).value)

        def f(col: int) -> float:
            return _safe_float(ws.cell(row, col).value, context=f"Centrales!{plant_name}, col {col}")

        def vol_or_from_cota(vol_col: int, cota_col: int) -> float:
            """Volumen columns (23-26) are themselves `=Vol_<Name>(cota)` formulas calling the
            FUNCCDEC_CDEC.xla add-in — when that add-in isn't loaded at recalc time Excel caches
            '#NAME?' instead of a number (confirmed: this happens in real, currently-used
            workbooks). Recover it ourselves using the exact same ported curve (see
            curves/reservoir_volume.py) against the paired Cota column (19-22), which isn't
            formula-derived and so doesn't have this failure mode."""
            raw = ws.cell(row, vol_col).value
            if isinstance(raw, (int, float)):
                return float(raw)
            if isinstance(raw, str) and raw.strip().upper() in _EXCEL_ERROR_STRINGS:
                cota = ws.cell(row, cota_col).value
                try:
                    return volume_from_level(plant_name, float(cota))
                except (TypeError, ValueError, KeyError) as exc:
                    print(
                        f"import_case: warning, Centrales!{plant_name} col {vol_col} is "
                        f"{raw!r} and recovering it from Cota (col {cota_col}={cota!r}) failed "
                        f"({exc}) — using 0.0."
                    )
                    return 0.0
            return _safe_float(raw, context=f"Centrales!{plant_name}, col {vol_col}")

        def rendimiento_or_from_cota() -> float:
            """Rendimiento (col 5) has the same '=Rend_<Name>(cota)' failure mode as the Volumen
            columns — see vol_or_from_cota above and curves/reservoir_yield.py. Only a subset of
            reservoirs have a ported Rend_<Name> curve at all (e.g. LMAULE genuinely has none in
            the source workbook), so this falls back to _safe_float's plain 0.0 default if there's
            no curve to recover from, same as any other unresolvable error."""
            raw = ws.cell(row, 5).value
            if isinstance(raw, (int, float)):
                return float(raw)
            if isinstance(raw, str) and raw.strip().upper() in _EXCEL_ERROR_STRINGS:
                cota = ws.cell(row, 19).value
                try:
                    return yield_from_level(plant_name, float(cota))
                except (TypeError, ValueError, KeyError) as exc:
                    print(
                        f"import_case: warning, Centrales!{plant_name} col 5 (rendimiento) is "
                        f"{raw!r} and recovering it from Cota (col 19={cota!r}) failed ({exc}) — "
                        "using 0.0."
                    )
                    return 0.0
            return _safe_float(raw, context=f"Centrales!{plant_name}, col 5")

        plant = Plant(
            case_id=case.id,
            cen_ind=int(cen_ind),
            name=plant_name,
            plant_type=plant_type,
            bus_id=bus.id if bus else None,
            cos_var=f(4),
            rendimiento=rendimiento_or_from_cota(),
            pot_min=f(27),
            pot_max=f(28),
        )
        if plant_type in ("EMBALSE", "SERIE"):
            plant.vert_min = f(29)
            plant.vert_max = f(30)
        if plant_type in ("EMBALSE", "SERIE", "PASADA"):
            plant.cau_afl = f(12)
            plant.estoc_findep = bool(ws.cell(row, 13).value)
        session.add(plant)
        session.flush()  # assign plant.id before Reservoir/self-FK resolution below

        if plant_type == "EMBALSE":
            session.add(
                Reservoir(
                    plant_id=plant.id,
                    vol_ini=vol_or_from_cota(23, 19),
                    vol_fin=vol_or_from_cota(24, 20),
                    vol_min=vol_or_from_cota(25, 21),
                    vol_max=vol_or_from_cota(26, 22),
                    f_esc=golden_femb.get(plant.name, 1.0),
                    cfue=bool(ws.cell(row, 9).value),
                )
            )

        plant_by_name[plant.name] = plant
        plant_by_cen_ind[plant.cen_ind] = plant
        gen_ref = ws.cell(row, 7).value
        vert_ref = ws.cell(row, 8).value
        pending_refs.append((plant, int(gen_ref) if gen_ref else 0, int(vert_ref) if vert_ref else 0))
        row += 1

    session.flush()
    for plant, gen_ref, vert_ref in pending_refs:
        if gen_ref:
            target = plant_by_cen_ind.get(gen_ref)
            if target is None:
                print(
                    f"import_case: warning, {plant.name!r} references downstream-gen cen_ind "
                    f"{gen_ref} which isn't an imported (non-'X') plant — leaving unset."
                )
            else:
                plant.downstream_gen_plant_id = target.id
        if vert_ref:
            target = plant_by_cen_ind.get(vert_ref)
            if target is None:
                print(
                    f"import_case: warning, {plant.name!r} references downstream-vert cen_ind "
                    f"{vert_ref} which isn't an imported (non-'X') plant — leaving unset."
                )
            else:
                plant.downstream_vert_plant_id = target.id
    session.flush()

    return plant_by_name


def _import_reservoir_curves(
    session: Session, case: Case, plant_by_name: dict[str, Plant], dat_static_dir: Path | None
) -> None:
    """plpcenre.dat/plpcenpmax.dat: no confirmed Excel/VBA source at all (see module docstring) —
    bootstrapped from the golden static files when available; skipped entirely (empty tables) for
    an xlsm-alone import, since there is no plausible default for a per-reservoir piecewise curve."""
    if dat_static_dir is None:
        print("import_case: warning, no dat_static_dir given — skipping plpcenre.dat/plpcenpmax.dat curves.")
        return
    cenre_path = dat_static_dir / "plpcenre.dat"
    if not cenre_path.exists():
        print("import_case: warning, no plpcenre.dat golden file — skipping reservoir yield curves.")
        return
    cenre = parse_plpcenre(cenre_path.read_text(encoding="latin-1"))
    for r in cenre["reservoirs"]:
        curve = ReservoirYieldCurve(
            case_id=case.id,
            plant_id=plant_by_name[r["plant_name"]].id,
            reservoir_plant_id=plant_by_name[r["reservoir_name"]].id,
            avg_yield=r["avg_yield"],
        )
        session.add(curve)
        session.flush()
        for seg in r["segments"]:
            session.add(
                ReservoirYieldSegment(
                    curve_id=curve.id,
                    ind=seg["ind"],
                    volume=seg["volume"],
                    slope=seg["slope"],
                    constant=seg["constant"],
                    scale=seg["scale"],
                )
            )

    cenpmax_path = dat_static_dir / "plpcenpmax.dat"
    if not cenpmax_path.exists():
        print("import_case: warning, no plpcenpmax.dat golden file — skipping reservoir Pmax curves.")
        session.flush()
        return
    cenpmax = parse_plpcenpmax(cenpmax_path.read_text(encoding="latin-1"))
    for r in cenpmax["reservoirs"]:
        curve = ReservoirPmaxCurve(
            case_id=case.id,
            plant_id=plant_by_name[r["plant_name"]].id,
            reservoir_plant_id=plant_by_name[r["reservoir_name"]].id,
        )
        session.add(curve)
        session.flush()
        for seg in r["segments"]:
            session.add(
                ReservoirPmaxSegment(
                    curve_id=curve.id, volume=seg["volume"], slope=seg["slope"], constant=seg["constant"]
                )
            )
    session.flush()


def _import_batteries(
    session: Session, case: Case, bus_by_num: dict[int, Bus], plant_by_name: dict[str, Plant], wb
) -> None:
    """Baterias sheet: header row 6, data from row 7 (INDICE, BATERIAS, Conectada a la Barra,
    Rendimiento de Descarga, Capacidad Mínima/Máxima, ..., Central de Carga, Batería objetivo,
    Rendimiento de Carga). This case has exactly one injector per battery (col 9), but the schema
    (and plpcenbat.dat's own format) supports more — extend this loop if a future case needs it."""
    ws = wb["Baterias"]
    row = 7
    while True:
        bat_ind = ws.cell(row, 1).value
        name = ws.cell(row, 2).value
        if bat_ind is None or name is None:
            break
        plant = plant_by_name.get(str(name))
        if plant is None:
            row += 1
            continue  # no matching BATERIA-type row in Centrales — skip rather than guess
        battery = Battery(
            case_id=case.id,
            plant_id=plant.id,
            bat_ind=int(bat_ind),
            bus_id=bus_by_num[int(ws.cell(row, 3).value)].id,
            discharge_loss_factor=_safe_float(ws.cell(row, 4).value, context=f"Baterias!{name}, discharge_loss_factor"),
            capacity_min=_safe_float(ws.cell(row, 5).value, context=f"Baterias!{name}, capacity_min"),
            capacity_max=_safe_float(ws.cell(row, 6).value, context=f"Baterias!{name}, capacity_max"),
        )
        session.add(battery)
        session.flush()
        injector_name = ws.cell(row, 9).value
        if injector_name is not None:
            session.add(
                BatteryInjector(
                    battery_id=battery.id,
                    name=str(injector_name),
                    loss_factor=_safe_float(ws.cell(row, 11).value, context=f"Baterias!{name}, loss_factor"),
                )
            )
        row += 1
    session.flush()


def _import_demand_profiles(
    session: Session, case: Case, bus_by_upper_name: dict[str, Bus], wb
) -> None:
    """Demanda-R/L/LD sheets: header rows 4-6, data from row 7, 24 rows per bus (one per hour),
    139 buses in this case. Columns 3-50 are (month, day-type) pairs: column = 4*month - 2 +
    day_type, day_type 1=Domingo/2=Lunes/3=Sabado/4=Trabajo — matches Rutina04.DEMxBarra2's own
    `ICol = 4 * imes - 2 + IDia` exactly. Bulk-inserted (Core `insert()`, not one ORM object per
    row) since this is ~480k rows for this case."""
    rows: list[dict] = []
    for category, sheet_name in (("R", "Demanda-R"), ("L", "Demanda-L"), ("LD", "Demanda-LD")):
        ws = wb[sheet_name]
        bar_row = 7
        while ws.cell(bar_row, 1).value is not None:
            bus = bus_by_upper_name.get(str(ws.cell(bar_row, 1).value).upper())
            if bus is not None:
                for hour in range(1, 25):
                    row = bar_row + hour - 1
                    for month in range(1, 13):
                        for day_type in range(1, 5):
                            col = 4 * month - 2 + day_type
                            mw = ws.cell(row, col).value
                            # Fast path avoids building a context string on every one of ~480k
                            # cells; _safe_float (with a real error message) only runs when the
                            # cheap type check below doesn't already prove it's a plain number.
                            if isinstance(mw, (int, float)):
                                mw_value = float(mw)
                            else:
                                mw_value = _safe_float(
                                    mw, context=f"{sheet_name}!{bus.name}, month {month} day-type {day_type} hour {hour}"
                                )
                            rows.append(
                                {
                                    "case_id": case.id,
                                    "bus_id": bus.id,
                                    "category": category,
                                    "month": month,
                                    "day_type": day_type,
                                    "hour": hour,
                                    "mw": mw_value,
                                }
                            )
            bar_row += 24
    session.execute(insert(DemandProfile), rows)
    session.flush()


def _import_consumption_and_holidays(session: Session, case: Case, wb) -> None:
    """Consumo sheet: week table from row 5 (Año/Mes/Semana/Inicial/Final/Nº días/GWh-R/GWh-L/
    GWh-LD), holiday list in column K ('Festivos') from row 5 — a plain date list, independent
    length from the week table."""
    ws = wb["Consumo"]
    row = 5
    week_num = 0
    while ws.cell(row, 4).value is not None:
        week_num += 1
        session.add(
            ConsumptionWeek(
                case_id=case.id,
                week_num=week_num,
                start_date=ws.cell(row, 4).value.date(),
                num_days=int(ws.cell(row, 6).value),
                gwh_r=_safe_float(ws.cell(row, 7).value, context=f"Consumo week {week_num}, gwh_r"),
                gwh_l=_safe_float(ws.cell(row, 8).value, context=f"Consumo week {week_num}, gwh_l"),
                gwh_ld=_safe_float(ws.cell(row, 9).value, context=f"Consumo week {week_num}, gwh_ld"),
            )
        )
        row += 1
    row = 5
    while ws.cell(row, 11).value is not None:
        session.add(Holiday(case_id=case.id, date=ws.cell(row, 11).value.date()))
        row += 1
    session.flush()


def _import_industrial_projects(
    session: Session, case: Case, bus_by_upper_name: dict[str, Bus], wb
) -> None:
    """Proyectos sheet: header row 5 (Fecha Inicial/Fecha Final/BARRA/Demanda/DESCRIPCIÓN), data
    from row 6."""
    ws = wb["Proyectos"]
    row = 6
    while ws.cell(row, 1).value is not None:
        bus = bus_by_upper_name.get(str(ws.cell(row, 3).value or "").upper())
        if bus is not None:
            session.add(
                IndustrialProject(
                    case_id=case.id,
                    bus_id=bus.id,
                    start_date=ws.cell(row, 1).value.date(),
                    end_date=ws.cell(row, 2).value.date(),
                    demand_mw=_safe_float(ws.cell(row, 4).value, context=f"Proyectos row {row}, demand_mw"),
                    description=ws.cell(row, 5).value,
                )
            )
        row += 1
    session.flush()


def _import_thermal_cost_schedule(
    session: Session, case: Case, plant_by_name: dict[str, Plant], wb
) -> None:
    """CV_MP sheet, columns G-J (CENTRAL/INICIAL/FINAL/[US$/MWh]) from row 6 — see
    ThermalCostSchedule's docstring for why this table (not the sheet's other, unrelated NAME/
    DATE1/DATE2/CV table in columns B-E) is plpcosce.dat's source. Iterates to the sheet's actual
    max_row since this table's row count doesn't match the B-E table's (confirmed: 13,331 vs
    15,449 populated rows) — they are genuinely different lengths, not misaligned copies of the
    same data."""
    ws = wb["CV_MP"]
    for row in range(6, ws.max_row + 1):
        central = ws.cell(row, 7).value
        if central is None:
            continue
        plant = plant_by_name.get(str(central))
        if plant is None:
            continue
        session.add(
            ThermalCostSchedule(
                case_id=case.id,
                plant_id=plant.id,
                stage_start=int(ws.cell(row, 8).value),
                stage_end=int(ws.cell(row, 9).value),
                cost_var=_safe_float(ws.cell(row, 10).value, context=f"CV_MP!{central}, cost_var"),
            )
        )
    session.flush()


def _date_range_to_stage_range(stages: list[Stage], d_start, d_end) -> tuple[int, int] | None:
    """MantEMBh is the one Phase 4 table with no pre-merged stage-range companion — this converts
    its raw [d_start, d_end] date range into a [stage_start, stage_end] num_eta range by finding
    every Stage whose own date span overlaps it. Returns None if nothing overlaps (silently
    dropped by the caller, matching the other tables' "ranges need not cover every plant/period"
    tolerance)."""
    matching = [
        s.num_eta
        for s in stages
        if s.start_date <= d_end and d_start <= s.start_date + timedelta(days=s.duration // 24 - 1)
    ]
    if not matching:
        return None
    return min(matching), max(matching)


def _import_plant_maintenance(session: Session, case: Case, plant_by_name: dict[str, Plant], wb) -> None:
    """MantCEN sheet: pre-merged block-range table, columns I-M (CENTRAL/INICIAL/FINAL/MÍNIMA/
    MÁXIMA), data from row 5. ~238k populated rows — read via `iter_rows` (much faster than
    per-cell access at this scale) and bulk-inserted."""
    ws = wb["MantCEN"]
    rows = []
    for central, ini, fin, pmin, pmax in ws.iter_rows(
        min_row=5, min_col=9, max_col=13, values_only=True
    ):
        if central is None:
            continue
        plant = plant_by_name.get(str(central))
        if plant is None:
            continue
        rows.append(
            {
                "case_id": case.id,
                "plant_id": plant.id,
                "block_start": int(ini),
                "block_end": int(fin),
                "pot_min": float(pmin),
                "pot_max": float(pmax),
            }
        )
    session.execute(insert(PlantMaintenance), rows)
    session.flush()


def _import_line_maintenance(session: Session, case: Case, wb) -> None:
    """MantLIN sheet: pre-merged block-range table, columns I-N (LÍNEA/INICIAL/FINAL/A-B/B-A/
    OPERATIVA), data from row 6."""
    ws = wb["MantLIN"]
    line_by_name = {ln.name: ln for ln in session.scalars(select(Line).where(Line.case_id == case.id))}
    row = 6
    while ws.cell(row, 9).value is not None:
        name = str(ws.cell(row, 9).value)
        line = line_by_name.get(name)
        if line is not None:
            session.add(
                LineMaintenance(
                    case_id=case.id,
                    line_id=line.id,
                    block_start=int(ws.cell(row, 10).value),
                    block_end=int(ws.cell(row, 11).value),
                    capacity_ab=_safe_float(ws.cell(row, 12).value, context=f"MantLIN!{name}, capacity_ab"),
                    capacity_ba=_safe_float(ws.cell(row, 13).value, context=f"MantLIN!{name}, capacity_ba"),
                    operational=str(ws.cell(row, 14).value).upper() in ("TRUE", "VERDADERO"),
                )
            )
        row += 1
    session.flush()


def _import_reservoir_maintenance(
    session: Session, case: Case, plant_by_name: dict[str, Plant], wb
) -> None:
    """MantEMB sheet: pre-merged stage-range table, columns H-L (EMBALSE/INICIAL/FINAL/MÍNIMO/
    MÁXIMO) — already volume-valued (Hm3), data from row 6."""
    ws = wb["MantEMB"]
    row = 6
    while ws.cell(row, 8).value is not None:
        plant = plant_by_name.get(str(ws.cell(row, 8).value))
        if plant is not None:
            session.add(
                ReservoirMaintenance(
                    case_id=case.id,
                    plant_id=plant.id,
                    stage_start=int(ws.cell(row, 9).value),
                    stage_end=int(ws.cell(row, 10).value),
                    vol_min=_safe_float(ws.cell(row, 11).value, context=f"MantEMB!{plant.name}, vol_min"),
                    vol_max=_safe_float(ws.cell(row, 12).value, context=f"MantEMB!{plant.name}, vol_max"),
                )
            )
        row += 1
    session.flush()


def _import_reservoir_min_volume_slack(
    session: Session, case: Case, plant_by_name: dict[str, Plant], stages: list[Stage], wb
) -> None:
    """MantEMBh sheet: raw date-range table, columns B-F (EMBALSE/INICIAL/FINAL/COTA/COSTO), data
    from row 6 — the one Phase 4 table needing date->stage conversion (see
    _date_range_to_stage_range). `level_min` is stored in the sheet's own Cota units; converted to
    volume via the ported Vol_<Name> curves only at generation time (see plpminembh.dat's generator)."""
    ws = wb["MantEMBh"]
    row = 6
    while ws.cell(row, 2).value is not None:
        plant = plant_by_name.get(str(ws.cell(row, 2).value))
        d_start = ws.cell(row, 3).value
        d_end = ws.cell(row, 4).value
        if plant is not None and d_start is not None and d_end is not None:
            stage_range = _date_range_to_stage_range(stages, d_start.date(), d_end.date())
            if stage_range is not None:
                session.add(
                    ReservoirMinVolumeSlack(
                        case_id=case.id,
                        plant_id=plant.id,
                        stage_start=stage_range[0],
                        stage_end=stage_range[1],
                        level_min=_safe_float(ws.cell(row, 5).value, context=f"MantEMBh!{plant.name}, level_min"),
                        cost=_safe_float(ws.cell(row, 6).value, context=f"MantEMBh!{plant.name}, cost"),
                    )
                )
        row += 1
    session.flush()


def _import_battery_maintenance(session: Session, case: Case, dat_block_dependant_dir: Path | None) -> None:
    """plpmanbat.dat: no Excel source at all (see module docstring) — bootstrapped from the
    golden file when available (else the table is simply left empty). Per the user's 2026-08-30
    ruling (code is the rule over a mismatched sample filename), this reads the checked-in
    `plpmantbat.dat` sample but the corresponding generator writes it out as `plpmanbat.dat`,
    matching what genpdbaterias.f actually opens."""
    if dat_block_dependant_dir is None:
        return
    battery_by_name = {b.plant.name: b for b in session.scalars(select(Battery).where(Battery.case_id == case.id))}
    golden_path = dat_block_dependant_dir / "plpmantbat.dat"
    if not golden_path.exists():
        return
    golden = parse_plpmanbat(golden_path.read_text(encoding="latin-1"))
    for b in golden["batteries"]:
        battery = battery_by_name.get(b["name"])
        if battery is None:
            continue
        for row in b["data"]:
            session.add(
                BatteryMaintenance(
                    case_id=case.id,
                    battery_id=battery.id,
                    block_start=row["num_blo"],
                    block_end=row["num_blo"],
                    e_min=row["e_min"],
                    e_max=row["e_max"],
                )
            )
    session.flush()


def _import_hydrology(
    session: Session,
    case: Case,
    plant_by_name: dict[str, Plant],
    stage_by_num: dict[int, Stage],
    dat_block_dependant_dir: Path | None,
) -> None:
    """plpaflce.dat/plpidsim.dat/plpidape.dat/plpidap2.dat: no Excel derivation at all — see
    db/models.py's Phase 5 section docstring for why (VBA's own `Rnd` PRNG isn't reproducible).
    Bootstrapped from the golden files when available; Inflow is bulk-inserted (~40k rows for this
    case). With no dat_block_dependant_dir (xlsm-alone import), all four tables are simply left
    empty — there is nothing in the workbook to derive hydrology data from regardless."""
    if dat_block_dependant_dir is None:
        print("import_case: warning, no dat_block_dependant_dir given — hydrology tables left empty.")
        return
    aflce_path = dat_block_dependant_dir / "plpaflce.dat"
    if aflce_path.exists():
        aflce = parse_plpaflce(aflce_path.read_text(encoding="latin-1"))
        rows = []
        for p in aflce["plants"]:
            plant = plant_by_name.get(p["name"])
            if plant is None:
                continue
            for b in p["blocks"]:
                rows.append(
                    {
                        "case_id": case.id,
                        "plant_id": plant.id,
                        "num_blo": b["num_blo"],
                        "values": b["values"],
                    }
                )
        session.execute(insert(Inflow), rows)

    idsim_path = dat_block_dependant_dir / "plpidsim.dat"
    if idsim_path.exists():
        idsim = parse_plpidsim(idsim_path.read_text(encoding="latin-1"))
        for s in idsim["stages"]:
            stage = stage_by_num.get(s["num_eta"])
            if stage is not None:
                session.add(
                    HydrologyScenarioAssignment(
                        case_id=case.id, stage_id=stage.id, hydro_class_by_sim=s["hydro_class"]
                    )
                )

    idape_path = dat_block_dependant_dir / "plpidape.dat"
    if idape_path.exists():
        idape = parse_plpidape(idape_path.read_text(encoding="latin-1"))
        for sim_idx, stages in enumerate(idape["simulations"], start=1):
            for s in stages:
                stage = stage_by_num.get(s["num_eta"])
                if stage is not None:
                    session.add(
                        ApertureIndexSimulation(
                            case_id=case.id,
                            simulation_slot=sim_idx,
                            stage_id=stage.id,
                            apertures=s["apertures"],
                        )
                    )

    idap2_path = dat_block_dependant_dir / "plpidap2.dat"
    if idap2_path.exists():
        idap2 = parse_plpidap2(idap2_path.read_text(encoding="latin-1"))
        for s in idap2["stages"]:
            stage = stage_by_num.get(s["num_eta"])
            if stage is not None:
                session.add(
                    ApertureIndexAggregate(case_id=case.id, stage_id=stage.id, apertures=s["apertures"])
                )

    session.flush()


def _import_basin_conventions(
    session: Session, case: Case, plant_by_name: dict[str, Plant], dat_static_dir: Path | None
) -> None:
    """plpralco.dat/plpextrac.dat/plpfilemb.dat/plpvrebemb.dat: no Excel source at all (see
    db/models.py's Phase 6 section docstring) — bootstrapped from the golden files when available,
    matched to Plant by name. plpmaulen.dat/plplajam.dat: real sheets exist but are stored as a
    verbatim ordered line sequence instead (see BasinConventionLine's docstring for why). With no
    dat_static_dir (xlsm-alone import), every table here is simply left empty."""

    def _read(name: str) -> str | None:
        if dat_static_dir is None:
            return None
        path = dat_static_dir / name
        return path.read_text(encoding="latin-1") if path.exists() else None

    if (text := _read("plpralco.dat")) is not None:
        d = parse_plpralco(text)
        plant = plant_by_name.get(d["name"])
        if plant is not None:
            session.add(RalcoConvention(case_id=case.id, plant_id=plant.id, segments=d["segments"]))

    if (text := _read("plpextrac.dat")) is not None:
        for p in parse_plpextrac(text)["points"]:
            source = plant_by_name.get(p["source"])
            downstream = plant_by_name.get(p["downstream"])
            if source is not None and downstream is not None:
                session.add(
                    ExtractionPoint(
                        case_id=case.id,
                        source_plant_id=source.id,
                        downstream_plant_id=downstream.id,
                        max_extraction=p["max_extraction"],
                    )
                )

    if (text := _read("plpfilemb.dat")) is not None:
        for r in parse_plpfilemb(text)["reservoirs"]:
            plant = plant_by_name.get(r["name"])
            downstream = plant_by_name.get(r["downstream"])
            if plant is not None and downstream is not None:
                session.add(
                    ReservoirFiltration(
                        case_id=case.id,
                        plant_id=plant.id,
                        downstream_plant_id=downstream.id,
                        avg_filtration=r["avg_filtration"],
                        segments=r["segments"],
                    )
                )

    if (text := _read("plpvrebemb.dat")) is not None:
        for r in parse_plpvrebemb(text)["reservoirs"]:
            plant = plant_by_name.get(r["name"])
            if plant is not None:
                session.add(
                    ReservoirSpillVolume(
                        case_id=case.id,
                        plant_id=plant.id,
                        spill_volume=r["spill_volume"],
                        cost=r["cost"],
                    )
                )

    for convention, filename in (("MAULE", "plpmaulen.dat"), ("LAJA", "plplajam.dat")):
        if (text := _read(filename)) is not None:
            for i, line in enumerate(parse_lines_raw(text)):
                session.add(
                    BasinConventionLine(
                        case_id=case.id,
                        convention=convention,
                        line_order=i,
                        is_comment=line["is_comment"],
                        text=line["text"],
                    )
                )

    session.flush()
